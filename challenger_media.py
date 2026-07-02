"""챌린저골프웨어 자동 입력 엔진.

타겟(결과물, 1개) = 통합보고서(=최종 리포트)의 일간리포트(자사몰)·일간리포트(브랜드스토어).
소스 = 챌린저골프웨어 매체 리포트(검색광고·브랜드검색·네이버GFA·모비온·구글) + META raw.
META는 일간리포트(자사몰) 'SNS 광고' 블록에 들어감 → 최종 리포트 한 파일로 완성.

설계 원칙(인수인계서 공통)
- 입력 셀만 채움(노출/유입(클릭)/주문(구매)/매출/광고비). CTR·광고효율·매출비중·파워링크합계(BG)는 수식 → 손대지 않음.
- 차트/서식/수식 보존 위해 build_report.patch_zip 사용(셀 값만 패치).

광고비 규칙(검증됨)
- 파워링크·GFA·브랜드검색 = 소스 광고비 그대로
- 모비온 = 소스 '광고비(VAT포함)' 열 그대로
- 구글 = 소스 광고비 × 1.1 / 0.85
- META(SNS) = 지출합 × 1.1 / 0.85
"""
import datetime, re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as CI
from build_report import patch_zip

_EPOCH = datetime.date(1899, 12, 30)


def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, str):
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', v)
        if m: return datetime.date(*map(int, m.groups()))
    if isinstance(v, (int, float)) and v > 40000:
        try: return _EPOCH + datetime.timedelta(days=int(v))
        except Exception: return None
    return None


def _n(v):
    return v if isinstance(v, (int, float)) else 0


# ── 채움 스펙: (타겟시트, 타겟열{필드:열}, 소스시트, 소스열{필드:열}, 광고비배수) ──
FIELDS = ["노출", "유입", "주문", "매출", "광고비"]
SPECS = [
    # 자사몰
    ("일간리포트 (자사몰)", dict(노출="AB", 유입="AC", 주문="AE", 매출="AF", 광고비="AG"),
     "브랜드검색", dict(노출="C", 유입="D", 주문="H", 매출="I", 광고비="G"), 1.0),
    ("일간리포트 (자사몰)", dict(노출="AS", 유입="AT", 주문="AU", 매출="AV", 광고비="AW"),
     "검색광고", dict(노출="N", 유입="O", 주문="S", 매출="T", 광고비="R"), 1.0),     # 파워링크 PC = M블록
    ("일간리포트 (자사몰)", dict(노출="AZ", 유입="BA", 주문="BB", 매출="BC", 광고비="BD"),
     "검색광고", dict(노출="Y", 유입="Z", 주문="AD", 매출="AE", 광고비="AC"), 1.0),   # 파워링크 MO = X블록
    ("일간리포트 (자사몰)", dict(노출="BN", 유입="BO", 주문="BQ", 매출="BR", 광고비="BS"),
     "네이버GFA", dict(노출="Y", 유입="Z", 주문="AD", 매출="AE", 광고비="AC"), 1.0),  # GFA 자사몰 = X블록
    ("일간리포트 (자사몰)", dict(노출="CD", 유입="CE", 주문="CG", 매출="CH", 광고비="CI"),
     "모비온", dict(노출="C", 유입="D", 주문="I", 매출="J", 광고비="H"), 1.0),        # 광고비=VAT포함(H)
    ("일간리포트 (자사몰)", dict(노출="DG", 유입="DH", 주문="DJ", 매출="DK", 광고비="DL"),
     "구글", dict(노출="O", 유입="P", 주문="U", 매출="V", 광고비="S"), 1.1 / 0.85),
    # 브랜드스토어
    ("일간리포트 (브랜드스토어)", dict(노출="AG", 유입="AH", 주문="AJ", 매출="AK", 광고비="AL"),
     "네이버GFA", dict(노출="N", 유입="O", 주문="S", 매출="T", 광고비="R"), 1.0),     # GFA 브랜드스토어 = M블록
]

# META(SNS) → 일간리포트(자사몰) SNS 광고 블록 입력칸
META_SNS = dict(노출="BV", 유입="BW", 주문="BY", 매출="BZ", 광고비="CA")


def _src_rowmap_B(ws):  # 소스 매체 리포트: 날짜=B열, 데이터 8행~
    m = {}
    for r in range(8, ws.max_row + 1):
        d = to_date(ws.cell(r, 2).value)
        if d: m[d] = r
    return m


def _tgt_rowmap_A(ws):  # 타겟 일간리포트: 날짜=A열(시리얼), 데이터 5행~
    m = {}
    for r in range(5, ws.max_row + 1):
        d = to_date(ws.cell(r, 1).value)
        if d: m[d] = r
    return m


def _meta_cols(ws):
    h = {str(ws.cell(1, c).value).strip(): c
         for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}
    col = lambda *ns: next((h[n] for n in ns if n in h), None)
    return dict(일=col("일"), 이름=col("광고 이름"), 노출=col("노출"), 클릭=col("링크 클릭"),
                도달=col("도달"), 구매=col("구매"), 장바=col("장바구니에 담기"),
                매출=col("구매 전환값"), 지출=col("지출 금액 (KRW)"))


def meta_daily(ws, cm, d):
    t = dict(노출=0, 유입=0, 주문=0, 매출=0, 지출=0)
    for r in range(2, ws.max_row + 1):
        if cm["일"] and to_date(ws.cell(r, cm["일"]).value) != d: continue
        t["노출"] += _n(ws.cell(r, cm["노출"]).value)
        t["유입"] += _n(ws.cell(r, cm["클릭"]).value)
        t["주문"] += _n(ws.cell(r, cm["구매"]).value)
        t["매출"] += _n(ws.cell(r, cm["매출"]).value)
        t["지출"] += _n(ws.cell(r, cm["지출"]).value)
    t["광고비"] = round(t["지출"] * 1.1 / 0.85)
    return t


# ──────────────────────────── META 소재별 효율 (주차별·소재별) ────────────────────────────
# 채움 항목: 노출(F)·클릭(G)·도달(I)·전환(J)·장바구니(K)·빈도(L)·매출(N)·광고비(O).
# 도달(I)은 일별 합산, 빈도(L)=노출/도달. 주간 도달은 중복제거 값이라 정확값과 다를 수 있음.
# 블록 맨 마지막 SUM/합계 행은 F가 수식이라 통째로 보존됨(빈도 합계행도 수기 유지).
MC_COLS = dict(노출="F", 클릭="G", 도달="I", 전환="J", 장바구니="K", 빈도="L", 매출="N", 광고비="O")


def _parse_creative(name):
    """광고 이름 → (기획전_명, 번호). 예: '0624_ASC_코어플렉스_001'→('코어플렉스','001')."""
    s = str(name)
    if s.startswith("2026"):
        return (s.split("_", 1)[1] if "_" in s else s, None)
    p = s.split("_")
    if len(p) >= 4 and p[1] in ("ASC", "트래픽"):
        return ("_".join(p[2:-1]), p[-1])
    return (s, None)


def _parse_week(b):
    """'2026.06.22~06.28' → (date(2026,6,22), date(2026,6,28))."""
    m = re.search(r'(\d{4})\.(\d{1,2})\.(\d{1,2})\s*~\s*(\d{1,2})\.(\d{1,2})', str(b))
    if not m: return None
    y, m1, d1, m2, d2 = map(int, m.groups())
    return (datetime.date(y, m1, d1), datetime.date(y, m2, d2))


def fill_meta_creative(meta_path, tgt_ws, dates):
    """META 소재별 효율의 '대상 날짜가 속한 주차 블록'을 raw로 재합산해 채움.
    반환: (edits_dict, 매칭실패 소재 리스트)."""
    mws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
    cm = _meta_cols(mws)

    # 주차 블록 헤더(B열 날짜범위) 행들
    headers = [r for r in range(1, tgt_ws.max_row + 1)
               if _parse_week(tgt_ws.cell(r, 2).value)]
    edits, missing = {}, []
    weeks_done = set()

    for d in sorted(dates):
        # d가 속한 블록 찾기
        blk = None
        for i, hr in enumerate(headers):
            wr = _parse_week(tgt_ws.cell(hr, 2).value)
            if wr and wr[0] <= d <= wr[1]:
                end = (headers[i + 1] - 1) if i + 1 < len(headers) else tgt_ws.max_row
                blk = (hr, end, wr); break
        if not blk or blk[0] in weeks_done:
            continue
        weeks_done.add(blk[0])
        start, end, (ws_, we_) = blk

        # 주차 raw 합산(기획전,번호별)
        agg = {}
        for r in range(2, mws.max_row + 1):
            dd = to_date(mws.cell(r, cm["일"]).value)
            if not dd or not (ws_ <= dd <= we_): continue
            k = _parse_creative(mws.cell(r, cm["이름"]).value)
            a = agg.setdefault(k, dict(노출=0, 클릭=0, 도달=0, 전환=0, 장바구니=0, 매출=0, 지출=0))
            a["노출"] += _n(mws.cell(r, cm["노출"]).value)
            a["클릭"] += _n(mws.cell(r, cm["클릭"]).value)
            a["도달"] += _n(mws.cell(r, cm["도달"]).value) if cm["도달"] else 0
            a["전환"] += _n(mws.cell(r, cm["구매"]).value)
            a["장바구니"] += _n(mws.cell(r, cm["장바"]).value) if cm["장바"] else 0
            a["매출"] += _n(mws.cell(r, cm["매출"]).value)
            a["지출"] += _n(mws.cell(r, cm["지출"]).value)

        # 시트 소재 행 매핑(기획전_명 carry-forward, 합계행/빈행 제외). 키 정규화(대소문자·구분자).
        def _norm(camp, num):
            c = str(camp or "").lower().replace(" ", "").replace("_", "")
            return (c, str(num) if num is not None else None)
        sheet_rows = {}
        cur = None
        for r in range(start, end + 1):
            f = tgt_ws.cell(r, CI("F")).value
            if isinstance(f, str) and f.startswith("="): continue          # 합계행
            dval = tgt_ws.cell(r, CI("D")).value
            eval_ = tgt_ws.cell(r, CI("E")).value
            if dval: cur = str(dval)
            if f is None and dval is None and eval_ is None: continue
            sheet_rows[_norm(cur, eval_)] = r

        matched = set()
        for k, a in agg.items():
            r = sheet_rows.get(_norm(*k))
            if r is None:
                if round(a["지출"]):  # 지출 0은 무시
                    missing.append((k[0], k[1], round(a["지출"] * 1.1 / 0.85)))
                continue
            matched.add(k)
            vals = dict(노출=a["노출"], 클릭=a["클릭"], 도달=a["도달"], 전환=a["전환"],
                        장바구니=a["장바구니"], 매출=a["매출"],
                        빈도=round(a["노출"] / a["도달"], 2) if a["도달"] else 0,
                        광고비=round(a["지출"] * 1.1 / 0.85))
            for fld, colL in MC_COLS.items():
                edits[f"{colL}{r}"] = vals[fld]

    return edits, missing


def run(source_path, meta_path, tonghap_path, out_path, only_dates=None):
    """source(매체 리포트) + meta → tonghap(통합보고서) 일간리포트 채움. 결과 1개 파일.
    날짜는 파일 값 기반으로 자동 인식 → 월(6월/7월…) 무관하게 동작."""
    src = load_workbook(source_path, data_only=True)
    tgt = load_workbook(tonghap_path, data_only=True)
    tgt_rows = {"일간리포트 (자사몰)": _tgt_rowmap_A(tgt["일간리포트 (자사몰)"]),
                "일간리포트 (브랜드스토어)": _tgt_rowmap_A(tgt["일간리포트 (브랜드스토어)"])}
    src_rows = {s: _src_rowmap_B(src[s]) for s in set(sp[2] for sp in SPECS)}

    # 대상 날짜: 지정 없으면, 소스 매체 리포트에 노출>0이고 통합보고서에도 존재하는 모든 날짜
    # (월/연도 하드코딩 없음 → 7월·8월 파일도 그대로 적응)
    if only_dates:
        dates = sorted(only_dates)
    else:
        ck = src["검색광고"]; rm = src_rows["검색광고"]
        tgt_self = tgt_rows["일간리포트 (자사몰)"]
        dates = sorted(d for d, r in rm.items()
                       if d in tgt_self and _n(ck.cell(r, CI("N")).value))

    edits = {"일간리포트 (자사몰)": {}, "일간리포트 (브랜드스토어)": {}}
    filled = 0
    log = []
    meta_missing = []

    for d in dates:
        # 1) 매체 블록(소스 리포트 → 통합보고서)
        for tsheet, tcols, ssheet, scols, admul in SPECS:
            trow = tgt_rows[tsheet].get(d)
            srow = src_rows[ssheet].get(d)
            if trow is None or srow is None: continue
            sw = src[ssheet]
            for fld in FIELDS:
                v = sw.cell(srow, CI(scols[fld])).value
                if v is None or isinstance(v, str): continue
                if fld == "광고비" and admul != 1.0:
                    v = round(v * admul)
                edits[tsheet][f"{tcols[fld]}{trow}"] = v
                filled += 1
        # 2) META(SNS) 블록
        if meta_path:
            trow = tgt_rows["일간리포트 (자사몰)"].get(d)
            if trow is not None:
                mws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
                cm = _meta_cols(mws)
                md = meta_daily(mws, cm, d)
                for fld, col in META_SNS.items():
                    edits["일간리포트 (자사몰)"][f"{col}{trow}"] = md[fld]
                    filled += 1
        log.append(f"[입력] {d} 완료")

    # 3) META 소재별 효율(대상 날짜가 속한 주차 블록 재합산)
    if meta_path and "META 소재별 효율" in tgt.sheetnames:
        mc_edits, meta_missing = fill_meta_creative(meta_path, tgt["META 소재별 효율"], dates)
        if mc_edits:
            edits["META 소재별 효율"] = mc_edits
            filled += len(mc_edits)
        if meta_missing:
            log.append("[META소재별] 시트에 행 없는 신규/잔여 소재(수기 행 추가 검토): "
                       + ", ".join(f"{a} {b}(광고비 {c:,}원)" for a, b, c in meta_missing))

    edits = {s: e for s, e in edits.items() if e}
    patch_zip(tonghap_path, edits, out_path)
    print(f"채운 셀: {filled}개 / 저장: {out_path}")
    for l in log: print("  ", l)
    return meta_missing


if __name__ == "__main__":
    import sys
    od = {datetime.date.fromisoformat(sys.argv[5])} if len(sys.argv) > 5 else None
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], od)
