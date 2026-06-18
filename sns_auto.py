"""SNS 소재별 효율 시트 자동 입력 (메타 Raw Data -> 주차 블록).
차트/썸네일/코멘트 보존을 위해 셀 값만 원본 ZIP에 직접 패치."""
import re, datetime, sys
from openpyxl import load_workbook
from collections import defaultdict
from build_report import patch_zip

_LAB='<rPr><b/><sz val="9"/><color indexed="81"/><rFont val="돋움"/><family val="3"/><charset val="129"/></rPr>'
_NUM='<rPr><b/><sz val="9"/><color indexed="81"/><rFont val="Tahoma"/><family val="2"/></rPr>'
def memo_text(single, collection):
    s=str(int(single)) if single else ''
    c=str(int(collection)) if collection else ''
    return ('<text>'
            f'<r>{_LAB}<t>단일이미지</t></r>'
            f'<r>{_NUM}<t xml:space="preserve"> : {s}\r\n</t></r>'
            f'<r>{_LAB}<t>컬렉션광고</t></r>'
            f'<r>{_NUM}<t xml:space="preserve"> : {c}</t></r>'
            '</text>')

def parse_name(name):
    name = str(name).strip()
    if name.startswith("카탈로그_"):
        return ("카탈로그", name.split("_",1)[1])
    if name.endswith("_제품"): name = name[:-3]   # 제품 변형 -> 같은 소재로 합산
    parts = name.split("_")
    if len(parts) < 4: return None
    token, rest = parts[2], "_".join(parts[3:])
    sinjyu = "(신규)" in token
    base = token.replace("(신규)","")
    if base.startswith("네이버"):
        b = base.replace("네이버","")
        brand = b if b in ("쟈딕","듀퐁","브로이어") else b
        return (brand, token + "_" + rest)
    if base.startswith("듀퐁"):
        prod = base.replace("듀퐁","")
        brand = "듀퐁(신규)" if sinjyu else "듀퐁"
        return (brand, (prod + "_" + rest) if prod else rest)
    return (base, rest)

def metric_groups(meta_path):
    ws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
    g = defaultdict(lambda: dict(노출=0,클릭=0,도달=0,전환=0,장바구니=0,단일전환=0,컬렉션전환=0))
    end = None
    for r in range(2, ws.max_row+1):
        a = ws.cell(r,1).value
        if not a: continue
        key = parse_name(a)
        if key is None: continue
        f = lambda c: (ws.cell(r,c).value if isinstance(ws.cell(r,c).value,(int,float)) else 0)
        d = g[key]
        d["노출"]+=f(8); d["클릭"]+=f(9); d["도달"]+=f(10); d["전환"]+=f(11); d["장바구니"]+=f(12)
        # 전환 광고형식 분해: 단일이미지=전환(제품X), 컬렉션광고=전환_제품 + 트래픽
        nm = str(a)
        if "_트래픽_" in nm or nm.endswith("_제품"):
            d["컬렉션전환"]+=f(11)
        else:
            d["단일전환"]+=f(11)
        u = ws.cell(r,21).value
        ud = None
        if isinstance(u,datetime.datetime): ud=u.date()
        elif isinstance(u,str):
            mm=re.search(r'(\d{4})-(\d{2})-(\d{2})',u)
            if mm: ud=datetime.date(int(mm.group(1)),int(mm.group(2)),int(mm.group(3)))
        if ud: end = max(end,ud) if end else ud
    return g, end

def week_blocks(ws):
    """SNS 시트의 주차 블록: [(시작행, 종료행, 시작일, 종료일)]"""
    blocks=[]; starts=[]
    for r in range(1, ws.max_row+1):
        b=ws.cell(r,2).value
        m=re.search(r"(\d+)월\s*(\d+)일~(\d+)월\s*(\d+)일", str(b) if b else "")
        if m:
            sd=datetime.date(2026,int(m.group(1)),int(m.group(2)))
            ed=datetime.date(2026,int(m.group(3)),int(m.group(4)))
            starts.append((r,sd,ed))
    for i,(r,sd,ed) in enumerate(starts):
        # 종료행 = TOTAL행 직전 (다음 블록 시작 전 또는 시트 끝)
        end_r = ws.max_row
        for r2 in range(r, ws.max_row+1):
            if ws.cell(r2,3).value=="TOTAL": end_r=r2-1; break
        blocks.append((r,end_r,sd,ed))
    return blocks

def block_rowmap(ws, r0, r1):
    rows={}; cur=""
    for r in range(r0, r1+1):
        c=ws.cell(r,3).value
        if c: cur=c
        e=ws.cell(r,5).value
        if e and cur!="TOTAL":
            rows[(cur, str(e).strip().lower())]=r
    return rows

def run(meta_path, target_path, out_path, account_totals=None):
    g, end = metric_groups(meta_path)
    wsd = load_workbook(target_path, data_only=True)["SNS 소재별 효율"]
    blocks = week_blocks(wsd)
    blk = next((b for b in blocks if b[2] <= (end or datetime.date(2026,6,17)) <= b[3]), blocks[-1])
    r0,r1,sd,ed = blk
    print(f"대상 주차: {sd}~{ed} (행 {r0}~{r1}) / 메타 종료일 {end}")
    rowmap = block_rowmap(wsd, r0, r1)
    edits = {"SNS 소재별 효율": {}}; E = edits["SNS 소재별 효율"]
    comment_edits = {"SNS 소재별 효율": {}}; CE = comment_edits["SNS 소재별 효율"]
    filled=0; new=[]
    for key, m in g.items():
        brand, jp = key
        rk = (brand, jp.strip().lower())
        if rk not in rowmap:
            if m['노출'] or m['클릭'] or m['전환'] or m['장바구니']:
                new.append((key, m))  # 데이터 있는 신규만
            continue
        row = rowmap[rk]
        freq = round(m["노출"]/m["도달"], 2) if m["도달"] else 0
        E[f"F{row}"]=m["노출"]; E[f"G{row}"]=m["클릭"]; E[f"H{row}"]=m["도달"]
        E[f"I{row}"]=m["전환"]; E[f"J{row}"]=m["장바구니"]; E[f"K{row}"]=freq
        CE[f"I{row}"] = memo_text(m["단일전환"], m["컬렉션전환"])   # 전환 분해 메모
        filled += 6
    if account_totals:
        # 대상 주차 TOTAL 행(블록 종료행+1)에 메타 계정 총합 기입 (도달/빈도는 중복제거라 합산 불가)
        trow = r1 + 1
        cols = dict(노출='F',클릭='G',도달='H',전환='I',장바구니='J',빈도='K')
        for k,col in cols.items():
            if k in account_totals:
                E[f'{col}{trow}'] = account_totals[k]
        print(f'TOTAL행 {trow}에 메타 계정 총합 기입')
    memo_n = patch_zip(target_path, edits, out_path, comment_edits=comment_edits)
    print(f"채운 셀: {filled} ({filled//6}개 소재) / 메모 갱신: {memo_n}건 / 저장: {out_path}")
    print(f"\n--- 신규 소재 {len(new)}건 (수동 추가 필요) ---")
    for (brand,jp),m in new:
        freq = round(m['노출']/m['도달'],2) if m['도달'] else 0
        print(f"  [{brand}] {jp}  노출{m['노출']:.0f} 클릭{m['클릭']:.0f} 도달{m['도달']:.0f} 전환{m['전환']:.0f} 장바구니{m['장바구니']:.0f} 빈도{freq}")
    return new

if __name__=="__main__":
    run(sys.argv[1], sys.argv[2], sys.argv[3])
