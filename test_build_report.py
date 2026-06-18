"""build_report 엔진 합성 테스트.
코드가 기대하는 포맷대로 통합보고서(소스)+최종리포트(타겟)을 만들고,
run() 후 타겟 셀에 소스 값이 실제로 들어갔는지 검증한다."""
import datetime, os, tempfile
from openpyxl import Workbook, load_workbook
import build_report

tmp = tempfile.mkdtemp()
DATES = [datetime.date(2026, 6, 15), datetime.date(2026, 6, 16), datetime.date(2026, 6, 17)]

# ── 소스: 통합보고서 ─────────────────────────────────
src = Workbook()
src.remove(src.active)
ws = src.create_sheet("네이버_브랜드검색")
# 블록 제목(2행) + 헤더(3행). base열을 C(3)에서 시작
base = 3
ws.cell(2, base, "듀코몰 PC")
hdr = ["노출수", "클릭수", "전환수", "c", "c", "c", "매출", "광고비"]  # base..base+7
for i, h in enumerate(hdr):
    ws.cell(3, base + i, h)
# 날짜(B열 5행~) + 값
for ridx, d in enumerate(DATES):
    r = 5 + ridx
    ws.cell(r, 2, d)
    ws.cell(r, base + 0, 1000 + ridx)   # 노출
    ws.cell(r, base + 1, 100 + ridx)    # 클릭
    ws.cell(r, base + 2, 10 + ridx)     # 전환
    ws.cell(r, base + 6, 500000 + ridx) # 매출
    ws.cell(r, base + 7, 50000 + ridx)  # 광고비
src_path = os.path.join(tmp, "tonghap.xlsx")
src.save(src_path)

# ── 타겟: 최종리포트 (N검색 시트) ───────────────────────
tgt = Workbook()
tgt.remove(tgt.active)
wt = tgt.create_sheet("N검색")
anchor = 2  # B열
wt.cell(1, anchor, "듀코몰 PC")          # 제목(헤더행-1)
wt.cell(2, anchor, "요일")
wt.cell(2, anchor + 1, "날짜")
for ridx, d in enumerate(DATES):
    r = 3 + ridx
    wt.cell(r, anchor + 1, d)            # C열 날짜 (리터럴 → data_only로 읽힘)
    # 값 칸은 비워둠 (노출=D, 클릭=E, 전환=H, 매출=I, 광고비=J)
tgt_path = os.path.join(tmp, "report.xlsx")
tgt.save(tgt_path)

# ── 실행 ─────────────────────────────────
out_path = os.path.join(tmp, "out.xlsx")
build_report.run(src_path, tgt_path, out_path, verify=False)

# ── 검증 ─────────────────────────────────
res = load_workbook(out_path, data_only=True)["N검색"]
print("\n=== 결과 검증 (N검색) ===")
ok = True
for ridx, d in enumerate(DATES):
    r = 3 + ridx
    vals = {c: res.cell(r, col).value for c, col in
            [("노출", 4), ("클릭", 5), ("전환", 8), ("매출", 9), ("광고비", 10)]}
    print(f"  {d}: {vals}")
    if vals["노출"] != 1000 + ridx:
        ok = False
print("\n결과:", "✅ 값 정상 입력됨" if ok else "❌ 값이 안 들어감(빈 셀)")
