"""듀코몰 추가 기능 자동입력:
1) [페이스북] 일자별  <- META raw 해당일 전체 합계 (광고비 = 지출×1.1÷0.85)
2) [듀퐁소품] 4표: 브검/파워링크(통합) + META 일자별(소품, 광고비×1.1÷0.85) + 소재별효율 소품(주간)
3) GFA 브로이어 카탈로그(AD82) <- 통합 네이버 GFA 브로이어 카탈로그
헤더 이름 기반으로 컬럼을 찾아 '일' 컬럼 유무에 무관하게 동작. 차트보존 위해 patch_zip.
"""
import datetime, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from build_report import patch_zip
import sns_auto  # parse_name, week_blocks, block_rowmap 재사용

MARKUP = 1.1 / 0.85  # 부가세10% + 마크업15%


def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, str):
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', v)
        if m: return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _num(v):
    return v if isinstance(v, (int, float)) else 0


# ---------- META ----------
def _meta_cols(ws):
    h = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}
    def col(*names):
        for nm in names:
            if nm in h: return h[nm]
        return None
    return dict(이름=col("광고 이름", "광고이름"), 일=col("일"),
                노출=col("노출"), 클릭=col("링크 클릭", "링크클릭"), 도달=col("도달"),
                구매=col("구매"), 장바구니=col("장바구니에 담기", "장바구니"),
                지출=col("지출 금액 (KRW)", "지출금액"), 매출=col("구매 전환값", "구매전환값"))


def meta_daily(meta_path, date, only_soum=False):
    """해당일 META 합계. 광고비=지출×1.1÷0.85. only_soum=소품 소재만."""
    ws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
    cm = _meta_cols(ws)
    t = dict(노출수=0, 클릭수=0, 전환수=0, 매출=0, 지출=0)
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, cm["이름"]).value
        if not nm: continue
        if cm["일"] and to_date(ws.cell(r, cm["일"]).value) != date: continue
        if only_soum:
            key = sns_auto.parse_name(nm)
            if not (key and str(key[1]).startswith("소품")): continue
        t["노출수"] += _num(ws.cell(r, cm["노출"]).value)
        t["클릭수"] += _num(ws.cell(r, cm["클릭"]).value)
        t["전환수"] += _num(ws.cell(r, cm["구매"]).value)
        t["매출"] += _num(ws.cell(r, cm["매출"]).value)
        t["지출"] += _num(ws.cell(r, cm["지출"]).value)
    t["광고비"] = round(t["지출"] * MARKUP)
    return t


# ---------- 일자별 표(공통) ----------
def daily_cols(ws, hrow, anchor=2, span=14):
    """헤더행에서 날짜/노출수/클릭수/전환수/매출/광고비 컬럼 위치."""
    cols = {}
    for c in range(anchor, anchor + span):
        v = ws.cell(hrow, c).value
        if v is None: continue
        h = str(v).replace("\n", "").replace(" ", "")
        for keyname in ("날짜", "노출수", "클릭수", "전환수", "광고비"):
            if keyname in h and keyname not in cols:
                cols[keyname] = c
        if "매출" in h and "매출" not in cols:  # 'PC 매출','브로이어 매출' 포함
            cols["매출"] = c
    return cols


def fill_daily(ws, hrow, date, vals, E, anchor=2):
    """vals = {노출수,클릭수,전환수,매출,광고비}. 해당 날짜행을 찾아 입력."""
    cols = daily_cols(ws, hrow, anchor)
    if "날짜" not in cols: return 0
    dcol = cols["날짜"]
    trow = None
    for r in range(hrow + 1, hrow + 60):
        if to_date(ws.cell(r, dcol).value) == date:
            trow = r; break
    if trow is None: return 0
    n = 0
    for k in ("노출수", "클릭수", "전환수", "매출", "광고비"):
        if k in cols and k in vals and vals[k] is not None:
            E[f"{get_column_letter(cols[k])}{trow}"] = vals[k]; n += 1
    return n


def find_title_row(ws, substr, col=2, rmax=300):
    for r in range(1, rmax):
        v = ws.cell(r, col).value
        if v and substr in str(v): return r
    return None


# ---------- 통합 소스 블록(일별) ----------
def source_block_daily(src_ws, label_substr, date):
    """행2 라벨에 label_substr 포함하는 블록을 찾아, 그 블록 헤더(행3)로 해당일 지표 반환."""
    c0 = None
    for c in range(1, src_ws.max_column + 1):
        v = src_ws.cell(2, c).value
        if v and label_substr in str(v): c0 = c; break
    if c0 is None: return None
    # 다음 블록 시작 전까지가 이 블록
    c1 = src_ws.max_column + 1
    for c in range(c0 + 1, src_ws.max_column + 1):
        if src_ws.cell(2, c).value is not None: c1 = c; break
    cmap = {}
    for c in range(c0, c1):
        h = src_ws.cell(3, c).value
        if h: cmap[str(h).strip()] = c
    drow = None
    for r in range(4, src_ws.max_row + 1):
        if to_date(src_ws.cell(r, 2).value) == date: drow = r; break
    if drow is None: return None
    def g(name):
        return _num(src_ws.cell(drow, cmap[name]).value) if name in cmap else None
    return dict(노출수=g("노출수"), 클릭수=g("방문자수"), 전환수=g("전환수"),
                매출=g("매출"), 광고비=g("광고비"))


# ---------- 소재별효율 소품(주간) ----------
def _week_blocks_flex(ws):
    """주차 블록(유연): '06월 01일 ~ 06월 07일' 등 공백 허용. -> [(r0,r1,sd,ed)]"""
    starts = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        m = re.search(r"(\d+)월\s*(\d+)일\s*~\s*(\d+)월\s*(\d+)일", str(b) if b else "")
        if m:
            sd = datetime.date(2026, int(m.group(1)), int(m.group(2)))
            ed = datetime.date(2026, int(m.group(3)), int(m.group(4)))
            starts.append((r, sd, ed))
    blocks = []
    for i, (r, sd, ed) in enumerate(starts):
        r1 = starts[i + 1][0] - 1 if i + 1 < len(starts) else ws.max_row
        blocks.append((r, r1, sd, ed))
    return blocks


def fill_soum_sojae(ws, meta_path, date, E):
    g, _ = sns_auto.metric_groups(meta_path)
    g = {k: v for k, v in g.items() if str(k[1]).startswith("소품")}  # 소품만
    blocks = _week_blocks_flex(ws)
    blk = next((b for b in blocks if b[2] <= date <= b[3]), blocks[-1] if blocks else None)
    if blk is None: return 0, None, []
    r0, r1, sd, ed = blk
    rowmap = sns_auto.block_rowmap(ws, r0, r1)
    n = 0; newc = []
    for key, m in g.items():
        brand, jp = key
        rk = (brand, jp.strip().lower())
        if rk not in rowmap:
            if m["노출"] or m["클릭"] or m["전환"] or m["장바구니"]:
                newc.append((key, m))
            continue
        row = rowmap[rk]
        freq = round(m["노출"] / m["도달"], 2) if m["도달"] else 0
        E[f"F{row}"] = m["노출"]; E[f"G{row}"] = m["클릭"]; E[f"H{row}"] = m["도달"]
        E[f"I{row}"] = m["전환"]; E[f"J{row}"] = m["장바구니"]; E[f"K{row}"] = freq
        n += 6
    return n, (sd, ed), newc


def run(report_path, tonghap_path, meta_path, out_path, only_dates=None):
    """only_dates=날짜 집합이면 그 날짜들, None이면 META 파일의 모든 '일' 날짜를 채움."""
    rwb = load_workbook(report_path, data_only=True)
    twb = load_workbook(tonghap_path, data_only=True)
    # 대상 날짜 결정
    if only_dates:
        target_dates = sorted(only_dates)
    else:
        mws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
        mc = _meta_cols(mws)
        ds = set()
        for r in range(2, mws.max_row + 1):
            d = to_date(mws.cell(r, mc["일"]).value) if mc["일"] else None
            if d: ds.add(d)
        target_dates = sorted(ds)
    edits = {}
    log = []
    fb = rwb["페이스북"]; dp = rwb["듀퐁소품"]; gfa = rwb["GFA"]
    Ef = edits.setdefault("페이스북", {}); Ed = edits.setdefault("듀퐁소품", {}); Eg = edits.setdefault("GFA", {})
    fb_hr = (find_title_row(fb, "일자별") or 29) + (0 if daily_cols(fb, find_title_row(fb, "일자별") or 29).get("날짜") else 1)
    bk_hr = (find_title_row(dp, "브랜드검색 PC 일자별") or 80) + 1
    pr = (find_title_row(dp, "파워링크 PC 일자별") or 134) + 1
    mr = (find_title_row(dp, "META 일자별") or 186) + 1
    gr = (find_title_row(gfa, "브로이어 카탈로그", col=30) or 82) + 1

    for date in target_dates:
        # 1) 페이스북 (전체 META 일별)
        fill_daily(fb, fb_hr, date, meta_daily(meta_path, date, only_soum=False), Ef)
        # 2a 듀퐁소품 브검 PC(앵커B) / MO(앵커P=16)  ※ B30 [일자별 성과]는 수식이라 건드리지 않음
        bpc = source_block_daily(twb["네이버_브랜드검색 듀퐁 소품"], "브랜드검색_PC", date)
        if bpc: fill_daily(dp, bk_hr, date, bpc, Ed, anchor=2)
        bmo = source_block_daily(twb["네이버_브랜드검색 듀퐁 소품"], "브랜드검색_MO", date)
        if bmo: fill_daily(dp, bk_hr, date, bmo, Ed, anchor=16)
        # 2b 듀퐁소품 파워링크 PC / MO
        ppc = source_block_daily(twb["네이버_파워링크 듀퐁 소품"], "파워링크_PC", date)
        if ppc: fill_daily(dp, pr, date, ppc, Ed, anchor=2)
        pmo = source_block_daily(twb["네이버_파워링크 듀퐁 소품"], "파워링크_MO", date)
        if pmo: fill_daily(dp, pr, date, pmo, Ed, anchor=16)
        # 2c 듀퐁소품 META 일자별 (소품)
        fill_daily(dp, mr, date, meta_daily(meta_path, date, only_soum=True), Ed)
        # 3) GFA 브로이어
        gk = source_block_daily(twb["네이버 GFA"], "브로이어 카탈로그", date)
        if gk: fill_daily(gfa, gr, date, gk, Eg, anchor=30)
    log.append(f"일별 입력 날짜: {[str(d) for d in target_dates]}")

    # 2d 듀퐁소품 소재별효율 소품 (대상 주차 1회)
    sn, rng, newc = fill_soum_sojae(dp, meta_path, max(target_dates), Ed)
    log.append(f"듀퐁소품 소재별효율 주차 {rng}: {sn//6}소재, 신규 {len(newc)}건")

    patch_zip(report_path, edits, out_path)
    print("\n".join(log))
    return newc


if __name__ == "__main__":
    import sys
    od = {datetime.date.fromisoformat(sys.argv[5])} if len(sys.argv) > 5 else None
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], od)
