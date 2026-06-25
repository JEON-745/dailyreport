"""NSR 최종 리포트 자동 입력 엔진 (단일 월시트 + 가로블록 구조).

대상(타겟): NSR 통합보고서의 월시트(예: '2026_6월')와 '메타 소재별 효율' 시트.
소스: 엠케이로드 통합보고서(매체별 시트) + META 소재별 raw('Raw Data Report').

설계 원칙(인수인계서 공통)
- 컬럼 위치는 '헤더 이름'으로 감지 (하드코딩 금지). 월/포맷이 달라져도 동작.
- 입력 셀만 채움: 노출수/유입자/주문수/(모든)주문수/매출액/광고비.
  CTR·광고효율·매출비중·전환율·ROAS·광고성과요약(AA~AE)은 전부 수식 -> 손대지 않음.
- 차트/서식/수식 보존 위해 build_report.patch_zip 사용(셀 값만 패치).

광고비 규칙
- 네이버(브검·파워링크·쇼핑검색)·GFA = raw 광고비 그대로
- 크리테오 = '광고비(vat 포함)' 열
- META(일별·소재별) = 지출 x 1.1 / 0.85  (부가세 10% + 마크업 15%)
"""
import datetime, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from build_report import patch_zip

MARKUP = 1.1 / 0.85


# ──────────────────────────── 공통 유틸 ────────────────────────────
_EXCEL_EPOCH = datetime.date(1899, 12, 30)  # 엑셀 날짜 시리얼 기준


def to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, str):
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', v)
        if m: return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    if isinstance(v, (int, float)) and v > 30000:   # 날짜 서식 셀이 시리얼 정수로 읽힘
        try:
            return _EXCEL_EPOCH + datetime.timedelta(days=int(v))
        except Exception:
            return None
    return None


def _num(v):
    return v if isinstance(v, (int, float)) else 0


def _norm(s):
    return str(s).replace("\n", "").replace(" ", "").strip() if s is not None else ""


# ──────────────────────────── 월시트(가로블록) ────────────────────────────
# 타겟 블록(3행 라벨) -> (소스시트, 소스2행 라벨 or None=단일블록, 소스 광고비 헤더)
MONTH_BLOCKS = [
    ("브랜드 검색 PC",     "클릭초이스_브검",     "브랜드검색_PC", "광고비"),
    ("브랜드 검색 MO",     "클릭초이스_브검",     "브랜드검색_MO", "광고비"),
    ("파워링크_PC",        "클릭초이스_파워링크", "파워링크_PC",   "광고비"),
    ("파워링크_MO",        "클릭초이스_파워링크", "파워링크_MO",   "광고비"),
    ("쇼핑검색_PC",        "클릭초이스_쇼핑검색", "쇼핑검색_PC",   "광고비"),
    ("쇼핑검색_MO",        "클릭초이스_쇼핑검색", "쇼핑검색_MO",   "광고비"),
    ("GFA 애드부스트 쇼핑", "GFA",                None,            "광고비"),
    ("리타겟팅",           "크리테오",           None,            "광고비(vat 포함)"),
]

# 타겟 입력열(4행 헤더) 정규화 이름 -> 내부 필드
TGT_FIELD = {
    "노출수": "노출수", "유입자": "유입자",
    "주문수": "주문수", "(모든)주문수": "모든주문수",
    "매출액": "매출액", "광고비": "광고비",
}


def detect_month_target(ws, row3=3, row4=4):
    """월시트 매체블록: 3행 라벨로 블록 시작열을 찾고, 4행 헤더로 입력열을 잡는다."""
    # 3행에서 매체 라벨이 있는 열 = 블록 시작점. 다음 라벨 전까지가 그 블록.
    label_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row3, c).value
        if v is not None and str(v).strip() != "":
            label_cols.append((c, str(v).strip()))
    blocks = {}
    for i, (c0, label) in enumerate(label_cols):
        c1 = label_cols[i + 1][0] if i + 1 < len(label_cols) else ws.max_column + 1
        cols = {}
        for c in range(c0, c1):
            h = _norm(ws.cell(row4, c).value)
            if not h: continue
            if "모든" in h and "주문" in h:
                cols.setdefault("모든주문수", c)
            elif h == "주문수":
                cols.setdefault("주문수", c)
            elif h in TGT_FIELD:
                cols.setdefault(TGT_FIELD[h], c)
        blocks[label] = cols
    return blocks


def month_date_row(ws, date, acol=1, r0=5):
    for r in range(r0, ws.max_row + 1):
        if to_date(ws.cell(r, acol).value) == date:
            return r
    return None


def detect_source_block(ws, label, adcol_hint="광고비"):
    """소스 시트에서 블록(2행 라벨=label, label=None이면 시트 전체)을 찾아
    3행 헤더로 컬럼맵 반환. 날짜는 B열(2), 데이터 5행~."""
    if label is None:
        c0, c1 = 1, ws.max_column + 1
    else:
        c0 = None
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(2, c).value) == _norm(label):
                c0 = c; break
        if c0 is None: return None
        c1 = ws.max_column + 1
        for c in range(c0 + 1, ws.max_column + 1):
            if ws.cell(2, c).value is not None:
                c1 = c; break
    cmap = {}
    for c in range(c0, c1):
        h = str(ws.cell(3, c).value).strip() if ws.cell(3, c).value is not None else ""
        if h and h not in cmap:
            cmap[h] = c
    return cmap


def src_field(cmap, *names):
    for nm in names:
        if nm in cmap: return cmap[nm]
    return None


def read_source_day(ws, cmap, date, ad_header):
    """해당 날짜행에서 입력 5~6개 값을 읽어 dict 반환. (B열 날짜)"""
    drow = None
    for r in range(5, ws.max_row + 1):
        if to_date(ws.cell(r, 2).value) == date:
            drow = r; break
    if drow is None: return None
    g = lambda col: ws.cell(drow, col).value if col else None
    c_노출  = src_field(cmap, "노출수")
    c_방문  = src_field(cmap, "방문자수")
    c_전환  = src_field(cmap, "전환수(기존)", "전환수(구매)", "전환수")
    c_모든  = src_field(cmap, "전환수(모든)")
    c_매출  = src_field(cmap, "매출")
    c_광고  = src_field(cmap, ad_header, "광고비")
    out = {
        "노출수":   g(c_노출),
        "유입자":   g(c_방문),
        "주문수":   g(c_전환),
        "모든주문수": g(c_모든),
        "매출액":   g(c_매출),
        "광고비":   g(c_광고),
    }
    return out


# ──────────────────────────── META raw ────────────────────────────
def _meta_cols(ws):
    h = {str(ws.cell(1, c).value).strip(): c
         for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}
    col = lambda *ns: next((h[n] for n in ns if n in h), None)
    return dict(이름=col("광고 이름", "광고이름"), 일=col("일"),
                노출=col("노출"), 클릭=col("링크 클릭", "링크클릭"), 도달=col("도달"),
                구매=col("구매"), 장바구니=col("장바구니에 담기", "장바구니"),
                매출=col("구매 전환값", "구매전환값"),
                지출=col("지출 금액 (KRW)", "지출금액"),
                종료=col("보고 종료", "보고종료"))


def meta_daily(ws, cm, date):
    """META 일별 합계. 광고비=지출x1.1/0.85, (모든)주문수=장바구니."""
    t = dict(노출수=0, 유입자=0, 주문수=0, 모든주문수=0, 매출액=0, 지출=0)
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, cm["이름"]).value: continue
        if cm["일"] and to_date(ws.cell(r, cm["일"]).value) != date: continue
        t["노출수"]    += _num(ws.cell(r, cm["노출"]).value)
        t["유입자"]    += _num(ws.cell(r, cm["클릭"]).value)
        t["주문수"]    += _num(ws.cell(r, cm["구매"]).value)
        t["모든주문수"] += _num(ws.cell(r, cm["장바구니"]).value)
        t["매출액"]    += _num(ws.cell(r, cm["매출"]).value)
        t["지출"]      += _num(ws.cell(r, cm["지출"]).value)
    t["광고비"] = round(t["지출"] * MARKUP, 2)
    return t


def nsr_creative_key(name):
    """'날짜범위_유형_기획전명' -> 기획전명. ASC/전환(링크최대화) 모두 같은 소재로 합산."""
    parts = str(name).strip().split("_")
    if len(parts) < 3: return None
    return "_".join(parts[2:])


# ──────────────────────────── 소재별 효율(주간) ────────────────────────────
def week_blocks(ws):
    """'06월 22일~06월 28일' 패턴으로 주차 블록 -> [(헤더행, 시작일, 종료일, r0, r1)]."""
    starts = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        m = re.search(r"(\d+)월\s*(\d+)일\s*~\s*(\d+)월\s*(\d+)일", str(b) if b else "")
        if m:
            sd = datetime.date(2026, int(m.group(1)), int(m.group(2)))
            ed = datetime.date(2026, int(m.group(3)), int(m.group(4)))
            starts.append((r, sd, ed))
    blocks = []
    for i, (r, sd, ed) in enumerate(starts):
        r1 = starts[i + 1][0] - 1 if i + 1 < len(starts) else ws.max_row
        blocks.append((r, sd, ed, r, r1))
    return blocks


def soso_rowmap(ws, r0, r1):
    """블록 내 소재명(E열) -> 행. TOTAL 행(E 비고 D·E 없음) 제외."""
    rows = {}
    for r in range(r0, r1 + 1):
        e = ws.cell(r, 5).value  # E 기획전 명
        if e is not None and str(e).strip():
            rows[str(e).strip()] = r
    return rows


def meta_creative_groups(ws, cm, upto_date=None):
    """raw를 기획전명 기준으로 합산(주차 누적). upto_date 지정 시 그 날짜까지만."""
    from collections import defaultdict
    g = defaultdict(lambda: dict(노출=0, 클릭=0, 도달=0, 전환=0, 장바구니=0, 매출=0, 지출=0))
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, cm["이름"]).value
        if not nm: continue
        if upto_date and cm["일"] and to_date(ws.cell(r, cm["일"]).value) and \
           to_date(ws.cell(r, cm["일"]).value) > upto_date:
            continue
        key = nsr_creative_key(nm)
        if not key: continue
        d = g[key]
        d["노출"]    += _num(ws.cell(r, cm["노출"]).value)
        d["클릭"]    += _num(ws.cell(r, cm["클릭"]).value)
        d["도달"]    += _num(ws.cell(r, cm["도달"]).value)
        d["전환"]    += _num(ws.cell(r, cm["구매"]).value)
        d["장바구니"] += _num(ws.cell(r, cm["장바구니"]).value)
        d["매출"]    += _num(ws.cell(r, cm["매출"]).value)
        d["지출"]    += _num(ws.cell(r, cm["지출"]).value)
    return g


# ──────────────────────────── 실행 ────────────────────────────
def run(tonghap_path, meta_path, target_path, out_path, only_dates=None, month_sheet=None):
    src = load_workbook(tonghap_path, data_only=True)
    meta_ws = load_workbook(meta_path, data_only=True)["Raw Data Report"]
    cm = _meta_cols(meta_ws)
    twb = load_workbook(target_path, data_only=True)

    # 월시트 자동 선택(없으면 첫 '2026_' 시트)
    if month_sheet is None:
        cand = [s for s in twb.sheetnames if re.match(r"\d{4}_\d+월", s)]
        month_sheet = cand[0] if cand else twb.sheetnames[0]
    mws = twb[month_sheet]

    # 대상 날짜
    if only_dates:
        target_dates = sorted(only_dates)
    else:
        ds = set()
        for r in range(2, meta_ws.max_row + 1):
            d = to_date(meta_ws.cell(r, cm["일"]).value) if cm["일"] else None
            if d: ds.add(d)
        target_dates = sorted(ds)

    edits = {month_sheet: {}, "메타 소재별 효율": {}}
    Em = edits[month_sheet]
    Es = edits["메타 소재별 효율"]
    log, filled = [], 0

    tgt_blocks = detect_month_target(mws)

    # ── 1) 월시트 매체블록(브검/파워/쇼핑/GFA/크리테오) ──
    for date in target_dates:
        trow = month_date_row(mws, date)
        if trow is None:
            log.append(f"[월시트] {date} 행 없음 -> skip"); continue
        for label, ss, src_label, ad_h in MONTH_BLOCKS:
            tcols = tgt_blocks.get(label)
            if not tcols: continue
            if ss not in src.sheetnames: continue
            cmap = detect_source_block(src[ss], src_label, ad_h)
            if not cmap: continue
            vals = read_source_day(src[ss], cmap, date, ad_h)
            if not vals: continue
            for fld, col in tcols.items():
                v = vals.get(fld)
                if v is None or isinstance(v, str): continue   # '-' 등 skip
                Em[f"{get_column_letter(col)}{trow}"] = v
                filled += 1

        # ── META 일별 ──
        mt = tgt_blocks.get("SNS 광고")
        if mt:
            md = meta_daily(meta_ws, cm, date)
            for fld, col in mt.items():
                v = md.get("광고비") if fld == "광고비" else md.get(fld)
                if v is None: continue
                Em[f"{get_column_letter(col)}{trow}"] = v
                filled += 1
        log.append(f"[월시트] {date} (행{trow}) 입력 완료")

    # ── 2) 메타 소재별 효율(주차 누적) ──
    swb_data = twb["메타 소재별 효율"]
    blocks = week_blocks(swb_data)
    fill_date = max(target_dates)
    blk = next((b for b in blocks if b[1] <= fill_date <= b[2]),
               blocks[-1] if blocks else None)
    new_creatives = []
    if blk:
        hr, sd, ed, r0, r1 = blk
        rowmap = soso_rowmap(swb_data, r0, r1)
        groups = meta_creative_groups(meta_ws, cm, upto_date=ed)
        # 헤더행 감지: 주차라벨행(hr)은 보통 첫 데이터행이라 헤더는 그 위에 있음.
        # hr에서 위로 올라가며 '노출수'가 있는 행을 헤더행으로 사용.
        header_row = None
        for rr in range(hr, max(hr - 3, 0), -1):
            if any(_norm(swb_data.cell(rr, c).value) == "노출수"
                   for c in range(2, swb_data.max_column + 1)):
                header_row = rr; break
        header_row = header_row or (hr - 1)
        hcols = {}
        for c in range(2, swb_data.max_column + 1):
            h = _norm(swb_data.cell(header_row, c).value)
            for k in ("노출수", "클릭수", "도달수", "전환", "장바구니", "빈도", "매출", "광고비"):
                if h == k and k not in hcols:
                    hcols[k] = c
        n_soso = 0
        for key, m in groups.items():
            if key not in rowmap:
                if any(m[x] for x in ("노출", "클릭", "전환", "장바구니")):
                    new_creatives.append((key, m))
                continue
            row = rowmap[key]
            freq = round(m["노출"] / m["도달"], 2) if m["도달"] else 0
            put = {
                "노출수": m["노출"], "클릭수": m["클릭"], "도달수": m["도달"],
                "전환": m["전환"], "장바구니": m["장바구니"], "빈도": freq,
                "매출": m["매출"], "광고비": round(m["지출"] * MARKUP, 2),
            }
            for k, c in hcols.items():
                if k in put:
                    Es[f"{get_column_letter(c)}{row}"] = put[k]; filled += 1
            n_soso += 1
        log.append(f"[소재별] 주차 {sd}~{ed} (행{r0}~{r1}): {n_soso}소재 입력, 신규 {len(new_creatives)}건")

    patch_zip(target_path, edits, out_path)
    print(f"채운 셀: {filled}개 / 저장: {out_path}")
    for l in log: print(" ", l)
    if new_creatives:
        print("\n--- 신규 소재(시트에 행 없음, 수동 추가 필요) ---")
        for key, m in new_creatives:
            print(f"  {key}: 노출{m['노출']:.0f} 클릭{m['클릭']:.0f} 전환{m['전환']:.0f} 장바구니{m['장바구니']:.0f}")
    return new_creatives


if __name__ == "__main__":
    import sys
    od = {datetime.date.fromisoformat(sys.argv[5])} if len(sys.argv) > 5 else None
    run(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], od)
