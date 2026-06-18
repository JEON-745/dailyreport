"""sns_auto 엔진 합성 테스트."""
import datetime, os, tempfile
from openpyxl import Workbook, load_workbook
import sns_auto

tmp = tempfile.mkdtemp()

# ── 소스: 메타 Raw Data Report ───────────────────────
meta = Workbook(); meta.remove(meta.active)
ws = meta.create_sheet("Raw Data Report")
# 헤더(1행)
ws.cell(1, 1, "소재명")
for c, h in [(8,"노출"),(9,"클릭"),(10,"도달"),(11,"전환"),(12,"장바구니"),(21,"일자")]:
    ws.cell(1, c, h)
# 데이터: 듀코몰_06월_듀퐁가방_봄기획전 -> ("듀퐁","가방_봄기획전")
rows = [
    ("듀코몰_06월_듀퐁가방_봄기획전",       2000, 200, 1500, 20, 5),
    ("듀코몰_06월_듀퐁가방_봄기획전_제품",   1000, 100,  800, 10, 3),  # 같은 소재로 합산
]
for ridx,(nm,no,cl,do,jn,jang) in enumerate(rows):
    r = 2 + ridx
    ws.cell(r,1,nm); ws.cell(r,8,no); ws.cell(r,9,cl); ws.cell(r,10,do)
    ws.cell(r,11,jn); ws.cell(r,12,jang)
    ws.cell(r,21, datetime.datetime(2026,6,16))
meta_path = os.path.join(tmp,"meta.xlsx"); meta.save(meta_path)

# ── 타겟: SNS 소재별 효율 ────────────────────────────
tgt = Workbook(); tgt.remove(tgt.active)
wt = tgt.create_sheet("SNS 소재별 효율")
wt.cell(1,2,"6월 15일~6월 21일")   # 주차 블록 헤더 (B열)
wt.cell(2,3,"듀퐁")                 # C=브랜드 카테고리
wt.cell(2,5,"가방_봄기획전")        # E=기획전명
wt.cell(3,3,"TOTAL")               # TOTAL 행 -> 블록 종료
tgt_path = os.path.join(tmp,"report.xlsx"); tgt.save(tgt_path)

out_path = os.path.join(tmp,"out.xlsx")
new = sns_auto.run(meta_path, tgt_path, out_path)

# ── 검증 ─────────────────────────────────
res = load_workbook(out_path, data_only=True)["SNS 소재별 효율"]
vals = {c: res.cell(2,col).value for c,col in
        [("노출",6),("클릭",7),("도달",8),("전환",9),("장바구니",10),("빈도",11)]}
print("\n=== 결과 검증 (행2) ===")
print("  ", vals)
exp_노출 = 2000+1000
ok = vals["노출"] == exp_노출 and vals["전환"] == 30
print("\n결과:", f"✅ 값 정상 입력됨 (노출 합산 {exp_노출})" if ok else "❌ 값이 안 들어감")
print("신규(수동추가) 목록:", new)
