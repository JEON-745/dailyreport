"""
엠케이로드 통합보고서 -> 듀코몰 최종 리포트 매체시트 자동 입력 엔진.

설계 원칙
- 소스/타겟 모두 '블록'을 자동 탐지 (행/열 하드코딩 최소화) -> 월이 바뀌어 행 위치가 달라져도 동작
- 블록 매칭 키 = (브랜드, 디바이스 PC/MO)
- 날짜 기준으로 5개 입력값만 복사: 노출수 / 클릭수(=방문자수) / 전환수 / 매출 / 광고비
- 광고비는 '광고비*부가세' 열이 있으면 그 값(VAT포함), 없으면 '광고비' 사용
"""
import sys, datetime, zipfile, re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 소스(통합보고서) 시트  ->  타겟(최종리포트) 시트
PAIRS = [
    (["모비온"],                                   "모비온"),
    (["네이버_브랜드검색"],                          "N검색"),
    (["네이버_파워링크"],                            "N파워링크"),
    (["네이버_쇼핑검색_듀퐁 가방",
      "네이버_쇼핑검색_듀퐁 셔츠"],                   "쇼핑검색"),
    (["네이버_쇼핑브랜드형_듀퐁 가방"],               "쇼핑브랜드형"),
    (["구글_키워드"],                                "구글_키워드"),
    (["네이버 GFA"],                                 "GFA"),   # 카탈로그(쟈딕/듀퐁)만 자동 매칭
]

# 광고비 추가 배수(매체 수수료 등). 구글은 통합보고서 부가세포함값 ÷ 0.85 (마크업 15%)
SPEND_MARKUP = {
    "구글_키워드": 1/0.85,
    "구글_GDN":   1/0.85,
}

# 구글_GDN: 통합보고서 구글_DA의 4개 캠페인 -> 최종 2개 표(좌/우)에 배치
#   (소스라벨 포함어, 소스제외어, 타겟표 제목 포함어, 좌B/우P)
GDN_ROUTE = [
    ("리타겟", "신규",  "리타겟",  "B"),
    ("신규",   None,    "리타겟",  "P"),
    ("스페셜", None,    "스페셜",  "B"),
    ("PMAX",   None,    "스페셜",  "P"),
]

# 특정 타겟시트는 제목에 해당 문자열이 든 표만 채움 (GFA=카탈로그만)
TGT_TITLE_REQUIRE = {
    "GFA": "카탈로그",
}

def norm_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    return None

def brand_key(title):
    """블록 제목 -> (브랜드, 디바이스) 정규화 키"""
    if title is None: return ("", "")
    t = str(title)
    dev = "MO" if "MO" in t else ("PC" if "PC" in t else "")
    if "듀코몰" in t: b = "듀코몰"
    elif "쟈딕" in t: b = "쟈딕"
    elif "브로이어" in t: b = "브로이어"
    elif "셔츠" in t: b = "듀퐁셔츠"
    elif "가방" in t: b = "듀퐁가방"
    elif "소품" in t: b = "듀퐁소품"
    elif "듀퐁" in t: b = "듀퐁"
    else: b = ""
    # 합계/총합계 블록은 무시
    if "총합계" in t or "합계" in t: return None
    return (b, dev)

# 소스 시트별 브랜드 강제지정(블록 제목에 제품명이 없는 경우)
SRC_BRAND_HINT = {
    "네이버_쇼핑검색_듀퐁 가방": "듀퐁가방",
    "네이버_쇼핑검색_듀퐁 셔츠": "듀퐁셔츠",
}

def detect_source_blocks(ws, brand_hint=None):
    """소스 시트의 블록들: 3행에서 '노출수' 위치를 찾고 2행 제목으로 키 생성"""
    blocks = []
    for c in range(2, ws.max_column + 1):
        if ws.cell(3, c).value == "노출수":
            title = ws.cell(2, c).value
            key = brand_key(title)
            if key is None:        # 합계 블록 skip
                continue
            if brand_hint:         # 제품명 강제지정
                key = (brand_hint, key[1])
            base = c
            vat = ws.cell(3, base + 8).value == "광고비*부가세"
            blocks.append({
                "key": key, "title": title,
                "col_노출": base, "col_클릭": base + 1, "col_전환": base + 2,
                "col_매출": base + 6,
                "col_광고비": base + (8 if vat else 7),
            })
    # 날짜->행 (B열 날짜, 5행부터)
    drow = {}
    for r in range(5, ws.max_row + 1):
        d = norm_date(ws.cell(r, 2).value)
        if d: drow[d] = r
    return blocks, drow

def detect_target_blocks(ws, wsd):
    """타겟 시트의 일자별 표들: '요일'+'날짜' 헤더를 찾아 블록화.
    ws=수식보존(쓰기용), wsd=계산값(날짜 읽기용)"""
    blocks = []
    for r in range(1, ws.max_row + 1):
        for anchor in (2, 16):  # B열 / P열
            if ws.cell(r, anchor).value == "요일" and ws.cell(r, anchor + 1).value == "날짜":
                title = ws.cell(r - 1, anchor).value
                key = brand_key(title) if title else ("", "")
                if key is None:
                    key = ("", "")
                # 날짜->행 (계산값 시트에서 읽기)
                drow = {}
                rr = r + 1
                blanks = 0
                while rr <= ws.max_row and rr <= r + 45:
                    d = norm_date(wsd.cell(rr, anchor + 1).value)
                    if d is None:
                        blanks += 1
                        if blanks >= 3: break
                        rr += 1; continue
                    blanks = 0
                    drow[d] = rr
                    rr += 1
                blocks.append({
                    "key": key, "title": title, "anchor": anchor, "drow": drow, "hr": r,
                    "col_노출": anchor + 2, "col_클릭": anchor + 3, "col_전환": anchor + 6,
                    "col_매출": anchor + 7, "col_광고비": anchor + 8,
                })
    return blocks

FIELDS = [("노출", "col_노출"), ("클릭", "col_클릭"), ("전환", "col_전환"),
          ("매출", "col_매출"), ("광고비", "col_광고비")]

def write_block(tws, tb, sb, sws, sdr, tgt_sheet, verify, mismatches, edits, only_dates=None):
    """소스블록 sb -> 타겟블록 tb 로 날짜 기준 5개 입력값을 edits에 기록(셀 직접쓰기X)"""
    n = 0
    sheet_edits = edits.setdefault(tgt_sheet, {})
    for d, trow in tb["drow"].items():
        if only_dates is not None and d not in only_dates: continue
        if d not in sdr: continue
        srow = sdr[d]
        for _, ck in FIELDS:
            val = sws.cell(srow, sb[ck]).value
            if isinstance(val, str): continue  # '-' 등 무효값 skip
            if val is None: continue
            if ck == "col_광고비" and tgt_sheet in SPEND_MARKUP:
                val = val * SPEND_MARKUP[tgt_sheet]
            ref = get_column_letter(tb[ck]) + str(trow)
            if verify:
                old = tws.cell(trow, tb[ck]).value
                if old is not None and abs((old or 0) - (val or 0)) > 0.5:
                    mismatches.append((tgt_sheet, tb["title"], d, ck, ref, old, val))
            sheet_edits[ref] = val
            n += 1
    return n

XL_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

def _join_pkg(base_dir, target):
    """OPC 관계의 Target을 패키지 루트 기준 경로로 정규화.
    절대('/xl/..'), 상대('../comments1.xml', 'worksheets/sheet1.xml') 모두 처리."""
    t = target.replace("\\", "/")
    if t.startswith("/"):
        return t.lstrip("/")
    out = []
    for p in (base_dir.rstrip("/") + "/" + t).split("/"):
        if p == "..":
            if out: out.pop()
        elif p in ("", "."):
            continue
        else:
            out.append(p)
    return "/".join(out)

def _col_idx(ref):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    col, row = m.group(1), int(m.group(2))
    n = 0
    for ch in col: n = n * 26 + (ord(ch) - 64)
    return n, row, col

def _fmt(v):
    if isinstance(v, bool): return "1" if v else "0"
    if isinstance(v, int): return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v)

def patch_zip(src_path, edits, out_path, comment_edits=None):
    """원본 xlsx의 해당 시트 XML에서 '셀 값'만 교체하고 나머지(차트/서식/도형)는 원본 그대로 복사.
    comment_edits={시트:{셀ref: '<text>...</text>'}} 가 있으면 해당 메모(코멘트) 텍스트도 교체.
    openpyxl 전체 재저장 시 차트가 깨지는 문제를 피하기 위함."""
    from lxml import etree
    z = zipfile.ZipFile(src_path)
    wbxml = z.read("xl/workbook.xml").decode("utf-8")
    relsxml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    # 시트명->rId, rId->파일경로 매핑. 속성 순서/절대경로에 무관하게 태그 단위로 파싱
    # (Excel/openpyxl/타도구마다 Id·Target·r:id 속성 순서와 Target 형식이 달라 정규식 순서 가정은 깨짐)
    def _attrs(tag):
        return dict(re.findall(r'([\w:]+)\s*=\s*"([^"]*)"', tag))

    name2rid = {}
    for tag in re.findall(r'<sheet\b[^>]*?/?>', wbxml):
        a = _attrs(tag)
        rid = a.get("r:id") or a.get("id")
        if a.get("name") and rid:
            name2rid[a["name"]] = rid

    rid2target = {}
    for tag in re.findall(r'<Relationship\b[^>]*?/?>', relsxml):
        a = _attrs(tag)
        if a.get("Id") and a.get("Target"):
            rid2target[a["Id"]] = a["Target"]

    sheetfile = {nm: _join_pkg("xl", rid2target[rid]) for nm, rid in name2rid.items() if rid in rid2target}

    patched = {}
    nsm = {"a": XL_NS}
    for sheet, cells in edits.items():
        path = sheetfile.get(sheet)
        if not path: continue
        tree = etree.fromstring(z.read(path))
        data = tree.find("a:sheetData", nsm)
        rowmap = {int(r.get("r")): r for r in data.findall("a:row", nsm)}
        for ref, val in cells.items():
            cidx, rnum, col = _col_idx(ref)
            row = rowmap.get(rnum)
            if row is None:
                row = etree.SubElement(data, f"{{{XL_NS}}}row"); row.set("r", str(rnum))
                rowmap[rnum] = row
            c = next((x for x in row.findall("a:c", nsm) if x.get("r") == ref), None)
            if c is None:
                c = etree.Element(f"{{{XL_NS}}}c"); c.set("r", ref)
                inserted = False
                for x in row.findall("a:c", nsm):
                    if _col_idx(x.get("r"))[0] > cidx:
                        x.addprevious(c); inserted = True; break
                if not inserted: row.append(c)
            if "t" in c.attrib: del c.attrib["t"]
            fEl = c.find("a:f", nsm)
            if fEl is not None: c.remove(fEl)
            vEl = c.find("a:v", nsm)
            if vEl is None: vEl = etree.SubElement(c, f"{{{XL_NS}}}v")
            vEl.text = _fmt(val)
        patched[path] = etree.tostring(tree, xml_declaration=True,
                                       encoding="UTF-8", standalone=True)

    if "<calcPr" in wbxml:
        wbxml2 = re.sub(r'<calcPr([^>]*?)/>',
                        lambda m: '<calcPr' + re.sub(r'\sfullCalcOnLoad="[^"]*"', '', m.group(1)) + ' fullCalcOnLoad="1"/>',
                        wbxml, count=1)
    else:
        wbxml2 = wbxml.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')
    patched["xl/workbook.xml"] = wbxml2.encode("utf-8")

    # ---- 메모(코멘트) 텍스트 교체 ----
    memo_n = 0
    if comment_edits:
        for sheet, refmap in comment_edits.items():
            if not refmap: continue
            spath = sheetfile.get(sheet)
            if not spath: continue
            base = spath.split("/")[-1]
            srel = f"xl/worksheets/_rels/{base}.rels"
            try:
                srelxml = z.read(srel).decode("utf-8")
            except KeyError:
                continue
            cmt = re.search(r'Target="([^"]*comments\d*\.xml)"', srelxml)
            if not cmt: continue
            cpath = _join_pkg("xl/worksheets", cmt.group(1))  # 워크시트 _rels 기준 상대경로
            raw = patched.get(cpath)
            if raw is None: raw = z.read(cpath)
            cxml = raw.decode("utf-8") if isinstance(raw, bytes) else raw
            for ref, text_xml in refmap.items():
                # 해당 ref 코멘트의 <text>...</text> 만 교체
                pat = re.compile(r'(<comment ref="' + re.escape(ref) + r'"[^>]*>)<text>.*?</text>(</comment>)', re.S)
                new_cxml, k = pat.subn(lambda mm: mm.group(1) + text_xml + mm.group(2), cxml)
                if k:
                    cxml = new_cxml; memo_n += 1
            patched[cpath] = cxml.encode("utf-8")

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zo:
        for item in z.infolist():
            zo.writestr(item, patched.get(item.filename, z.read(item.filename)))
    z.close()
    return memo_n

def run(tonghap_path, target_path, out_path, verify=True, only_dates=None):
    src_wb = load_workbook(tonghap_path, data_only=True)
    tgt_wb = load_workbook(target_path)  # 수식 보존(쓰기용)
    tgt_data = load_workbook(target_path, data_only=True)  # 날짜 읽기용
    log, mismatches, filled = [], [], 0
    edits = {}  # {시트명: {셀주소: 값}}  -> 마지막에 원본 ZIP에 직접 패치

    for src_sheets, tgt_sheet in PAIRS:
        if tgt_sheet not in tgt_wb.sheetnames:
            log.append(f"[SKIP] 타겟시트 없음: {tgt_sheet}"); continue
        tws = tgt_wb[tgt_sheet]
        tblocks = detect_target_blocks(tws, tgt_data[tgt_sheet])

        # 소스 블록 모으기 (여러 소스시트 -> 한 타겟)
        sblocks, sdrow = [], None
        for ss in src_sheets:
            if ss not in src_wb.sheetnames:
                log.append(f"[WARN] 소스시트 없음: {ss}"); continue
            sws = src_wb[ss]
            b, dr = detect_source_blocks(sws, SRC_BRAND_HINT.get(ss))
            for blk in b: blk["_sws"] = sws; blk["_drow"] = dr
            sblocks += b

        # 타겟 블록 단위로 매칭
        req = TGT_TITLE_REQUIRE.get(tgt_sheet)
        for tb in tblocks:
            if req and (not tb["title"] or req not in str(tb["title"])):
                continue  # 지정 키워드 없는 표는 건너뜀
            # 1) (브랜드,디바이스) 정확 매칭
            cand = [s for s in sblocks if s["key"] == tb["key"]]
            # 2) 디바이스만 일치 (브랜드 라벨이 다른 경우: 쇼핑브랜드형 등)
            if not cand and tb["drow"]:
                same_dev = [s for s in sblocks if s["key"][1] == tb["key"][1]]
                if len(same_dev) == 1:
                    cand = same_dev
            # 3) 단일블록 시트(모비온 등)
            if not cand and tb["key"] == ("", "") and len(sblocks) == 1:
                cand = sblocks
            if not cand:
                if tb["drow"]:
                    log.append(f"[{tgt_sheet}] 매칭없음 타겟블록 {tb['key']} \"{tb['title']}\"")
                continue
            sb = cand[0]; sws = sb["_sws"]; sdr = sb["_drow"]
            filled += write_block(tws, tb, sb, sws, sdr, tgt_sheet, verify, mismatches, edits, only_dates)
            log.append(f"[{tgt_sheet}] OK {tb['key']} \"{tb['title']}\" <- \"{sb['title']}\" ({len(tb['drow'])}일)")

    # ---- 구글_GDN 전용: 구글_DA 4개 캠페인 -> 최종 2개 표(좌/우) ----
    if "구글_GDN" in tgt_wb.sheetnames and "구글_DA" in src_wb.sheetnames:
        tgt_sheet = "구글_GDN"
        tws = tgt_wb[tgt_sheet]
        tblocks = detect_target_blocks(tws, tgt_data[tgt_sheet])
        sws = src_wb["구글_DA"]
        sblocks, sdr = detect_source_blocks(sws)  # 4개 캠페인 블록(키는 모두 듀코몰,'')
        for blk in sblocks: blk["title_s"] = str(blk["title"])
        for src_kw, src_ex, tgt_kw, side in GDN_ROUTE:
            # 소스 블록 찾기
            sb = next((b for b in sblocks if src_kw in b["title_s"]
                       and (src_ex is None or src_ex not in b["title_s"])), None)
            if not sb: continue
            # 타겟 표 찾기: 제목에 tgt_kw 포함하는 B표의 헤더행 -> side에 맞는 anchor
            base_tb = next((t for t in tblocks if t["anchor"] == 2 and t["title"]
                            and tgt_kw in str(t["title"])), None)
            if not base_tb: continue
            want_anchor = 2 if side == "B" else 16
            tb = next((t for t in tblocks if t["hr"] == base_tb["hr"] and t["anchor"] == want_anchor), None)
            if not tb: continue
            n = write_block(tws, tb, sb, sws, sdr, tgt_sheet, verify, mismatches, edits, only_dates)
            filled += n
            log.append(f"[구글_GDN] OK {side} \"{tb['title'] or base_tb['title']}\" <- \"{sb['title']}\" ({n//5}일)")

    patch_zip(target_path, edits, out_path)
    print(f"\n채운 셀: {filled}개  / 저장: {out_path}")
    print("\n--- 매칭 로그 ---")
    for l in log: print(" ", l)
    if verify:
        print(f"\n--- 검증: 기존값과 불일치 {len(mismatches)}건 ---")
        for m in mismatches[:40]:
            print(f"  {m[0]} | {m[1]} | {m[2]} | {m[3]} | {m[4]} 기존={m[5]} 새값={m[6]}")
    return mismatches

if __name__ == "__main__":
    run(sys.argv[1], sys.argv[2], sys.argv[3])
